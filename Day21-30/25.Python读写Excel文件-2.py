"""
写Excel文件
"""

import random
import datetime
import openpyxl

# 第一步：创建工作簿（Workbook）
wb = openpyxl.Workbook()

# 第二步：添加工作表（Worksheet）
sheet = wb.active
sheet.title = "期末成绩"

titles = ("姓名", "语文", "数学", "英语")
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)

names = ("关羽", "张飞", "赵云", "马超", "黄忠")
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name)
    for col_index in range(2, 5):
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

# 第四步：保存工作簿
wb.save("考试成绩表.xlsx")


"""
读取Excel文件
"""


# 加载一个工作簿 ---> Workbook
wb = openpyxl.load_workbook("考试成绩表.xlsx")
# 获取工作表的名字
print(f"wb.sheetnames : {wb.sheetnames}")

# 获取工作表 ---> Worksheet
sheet = wb.worksheets[0]

# 获得单元格的范围
print(f"sheet.dimensions : {sheet.dimensions}")

# 获得行数和列数
print(f"行数: {sheet.max_row}, 列数: {sheet.max_column}")

# 获取指定单元格的值
print(f"第3行第3列的值 : {sheet.cell(3, 3).value}")
print(f"C3单元格的值 : {sheet['C3'].value}")

# 获取多个单元格（嵌套元组）
print(f"A2到C5单元格的值 : {sheet['A2:C5']}")

# 读取所有单元格的数据
for row_ch in range(2, sheet.max_row + 1):
    for col_ch in "ABCD":
        value = sheet[f"{col_ch}{row_ch}"].value
        if type(value) == datetime.datetime:
            print(value.strftime("%Y年%m月%d日"), end="\t")
        elif type(value) == int:
            print(f"{value:<10d}", end="\t")
        elif type(value) == float:
            print(f"{value:.4f}", end="\t")
        else:
            print(value, end="\t")
    print()


"""
调整样式和公式计算
"""
from openpyxl.styles import Font, Alignment, Border, Side

# 对齐方式
alignment = Alignment(horizontal="center", vertical="center")
# 边框线条
side = Side(color="ff7f50", style="mediumDashed")

wb = openpyxl.load_workbook("考试成绩表.xlsx")
sheet = wb.worksheets[0]

# 调整行高和列宽
sheet.row_dimensions[1].height = 30
sheet.column_dimensions["E"].width = 120

sheet["E1"] = "平均分"

# 设置字体
sheet.cell(1, 5).font = Font(size=18, bold=True, color="ff1493", name="华文楷体")
# 设置对齐方式
sheet.cell(1, 5).alignment = alignment
# 设置单元格边框
sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
for i in range(2, 7):
    # 公式计算每个学生的平均分
    sheet[f"E{i}"] = f"=average(B{i}:D{i})"
    sheet.cell(i, 5).font = Font(size=12, color="4169e1", italic=True)
    sheet.cell(i, 5).alignment = alignment

wb.save("考试成绩表_样式_公式.xlsx")


"""
生成统计图表
"""
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook(write_only=True)
sheet = wb.create_sheet()

rows = [
    ("类别", "销售A组", "销售B组"),
    ("手机", 40, 30),
    ("平板", 50, 60),
    ("笔记本", 80, 70),
    ("外围设备", 20, 10),
]

# 向表单中添加行
for row in rows:
    sheet.append(row)

# 创建图表对象
chart = BarChart()
chart.type = "col"
chart.style = 10
# 设置图表的标题
chart.title = "销售统计图"
# 设置图表纵轴的标题
chart.y_axis.title = "销量"
# 设置图表横轴的标题
chart.x_axis.title = "商品类别"
# 设置数据的范围
data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=3)
# 设置分类的范围
cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
# 给图表添加数据
chart.add_data(data, titles_from_data=True)
# 给图表设置分类
chart.set_categories(cats)
chart.shape = 4
# 将图表添加到表单指定的单元格中
sheet.add_chart(chart, "A10")

wb.save("demo.xlsx")
